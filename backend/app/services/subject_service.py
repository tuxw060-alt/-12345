"""Account subject business logic."""

import re
import uuid
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.account_subject import AccountSubject
from app.models.journal_entry import JournalEntryLine
from app.schemas.account_subject import (
    SubjectCreate,
    SubjectImportConflict,
    SubjectImportResponse,
    SubjectUpdate,
)


CURRENT_PARENT_META: dict[str, tuple[str, str]] = {
    "1122": ("资产", "debit"),
    "1123": ("资产", "debit"),
    "1221": ("资产", "debit"),
    "2202": ("负债", "credit"),
    "2203": ("负债", "credit"),
    "2241": ("负债", "credit"),
}
CURRENT_PARENT_NAMES: dict[str, str] = {
    "1122": "应收账款",
    "1123": "预付账款",
    "1221": "其他应收款",
    "2202": "应付账款",
    "2203": "预收账款",
    "2241": "其他应付款",
}
CURRENT_PARENT_CODES = tuple(CURRENT_PARENT_META.keys())
HEADER_WORDS = {"编码", "编号", "代码", "科目编码", "科目代码", "名称", "科目名称", "单位名称"}


@dataclass
class LegacySubjectRow:
    code: str
    name: str
    full_code: str
    parent_code: str
    parent_name: str


@dataclass
class LegacySubjectWorkbook:
    parent_code: str
    parent_name: str
    rows: list[LegacySubjectRow]
    warnings: list[str]


def normalize_party_name(value: str | None) -> str:
    return re.sub(r"\s+", "", value or "")


def is_current_parent_code(code: str | None) -> bool:
    return (code or "").strip() in CURRENT_PARENT_META


def is_current_subject_code(code: str | None) -> bool:
    return any((code or "").strip().startswith(parent) for parent in CURRENT_PARENT_CODES)


async def list_subjects(
    db: AsyncSession,
    client_id: str | None = None,
    category: str | None = None,
    search: str | None = None,
    leaf_only: bool = False,
    offset: int = 0,
    limit: int = 500,
):
    stmt = select(AccountSubject).where(AccountSubject.is_active == True)

    if client_id is not None:
        stmt = stmt.where(
            (AccountSubject.client_id == None) | (AccountSubject.client_id == client_id)
        )
    if category:
        stmt = stmt.where(AccountSubject.category == category)
    if search:
        stmt = stmt.where(
            (AccountSubject.code.contains(search))
            | (AccountSubject.name.contains(search))
            | (AccountSubject.full_name.contains(search))
        )
    if leaf_only:
        stmt = stmt.where(AccountSubject.is_leaf == True)

    total = (await db.execute(select(func.count()).select_from(stmt.subquery()))).scalar() or 0
    result = await db.execute(stmt.order_by(AccountSubject.code).offset(offset).limit(limit))
    return result.scalars().all(), total


async def get_subject_by_code(db: AsyncSession, code: str) -> AccountSubject | None:
    result = await db.execute(select(AccountSubject).where(AccountSubject.code == code))
    return result.scalar_one_or_none()


async def get_subject_by_id(db: AsyncSession, subject_id: str) -> AccountSubject | None:
    result = await db.execute(select(AccountSubject).where(AccountSubject.id == subject_id))
    return result.scalar_one_or_none()


async def create_subject(db: AsyncSession, data: SubjectCreate) -> AccountSubject:
    subject = AccountSubject(
        id=str(uuid.uuid4()),
        client_id=data.client_id,
        code=data.code,
        name=data.name,
        full_name=data.full_name or data.name,
        level=data.level,
        parent_code=data.parent_code,
        parent_account_name=data.parent_account_name,
        category=data.category,
        direction=data.direction,
        is_leaf=data.is_leaf,
        created_from=data.created_from,
    )
    db.add(subject)
    await db.flush()
    return subject


async def update_subject(
    db: AsyncSession, subject: AccountSubject, data: SubjectUpdate
) -> AccountSubject:
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(subject, field, value)
    await db.flush()
    return subject


async def build_subject_tree(
    db: AsyncSession, client_id: str | None = None
) -> list[dict]:
    stmt = select(AccountSubject).where(AccountSubject.is_active == True)
    if client_id is not None:
        stmt = stmt.where(
            (AccountSubject.client_id == None) | (AccountSubject.client_id == client_id)
        )
    result = await db.execute(stmt.order_by(AccountSubject.code))
    subjects = result.scalars().all()

    node_map: dict[str, dict] = {}
    roots: list[dict] = []
    for subject in subjects:
        node_map[subject.code] = {
            "code": subject.code,
            "name": subject.name,
            "full_name": subject.full_name,
            "level": subject.level,
            "parent_code": subject.parent_code,
            "parent_account_name": subject.parent_account_name,
            "direction": subject.direction,
            "is_leaf": subject.is_leaf,
            "children": [],
        }

    for subject in subjects:
        node = node_map[subject.code]
        if subject.parent_code and subject.parent_code in node_map:
            node_map[subject.parent_code]["children"].append(node)
        else:
            roots.append(node)
    return roots


def generate_next_sub_account_code(parent_code: str, existing_child_codes: list[str]) -> str:
    parent_code = str(parent_code).strip()
    suffixes: list[str] = []
    for code in existing_child_codes:
        code = str(code).strip()
        if code.startswith(parent_code) and len(code) > len(parent_code):
            suffix = code[len(parent_code):]
            if suffix.isdigit():
                suffixes.append(suffix)
    width = max((len(suffix) for suffix in suffixes), default=2)
    next_number = max((int(suffix) for suffix in suffixes), default=0) + 1
    return f"{parent_code}{next_number:0{width}d}"


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).replace("\u3000", " ").replace("\xa0", " ").strip()


def _scope_clause(client_id: str | None):
    return AccountSubject.client_id == client_id if client_id else AccountSubject.client_id == None


def _split_code_name(text: str) -> tuple[str, str] | None:
    match = re.match(r"^\s*([0-9][0-9A-Za-z.\-_]*)\s+(.+?)\s*$", text or "")
    if not match:
        return None
    return match.group(1).strip(), match.group(2).strip()


def _looks_like_header(code: str, name: str) -> bool:
    return code in HEADER_WORDS or name in HEADER_WORDS


def _category_direction(parent_code: str) -> tuple[str, str]:
    for prefix, meta in CURRENT_PARENT_META.items():
        if parent_code.startswith(prefix):
            return meta
    return "资产", "debit"


def _read_xls_rows(content: bytes) -> list[list[str]]:
    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    return [
        [_cell_text(sheet.cell_value(row_index, col)) for col in range(sheet.ncols)]
        for row_index in range(sheet.nrows)
    ]


def _read_xlsx_rows(content: bytes) -> list[list[str]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = [[_cell_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return rows


def _resolve_child_code(
    current_parent_code: str,
    current_parent_name: str,
    child_code: str,
) -> tuple[str, str, str]:
    child_code = child_code.strip()
    if len(child_code) > 4 and child_code[:4] in CURRENT_PARENT_META:
        parent_code = child_code[:4]
        parent_name = (
            current_parent_name
            if parent_code == current_parent_code and current_parent_name
            else CURRENT_PARENT_NAMES.get(parent_code, current_parent_name)
        )
        return child_code, parent_code, parent_name
    return f"{current_parent_code}{child_code}", current_parent_code, current_parent_name


def parse_legacy_subject_workbook(filename: str, content: bytes) -> LegacySubjectWorkbook:
    ext = Path(filename).suffix.lower()
    if ext == ".xls":
        rows = _read_xls_rows(content)
    elif ext in {".xlsx", ".xlsm"}:
        rows = _read_xlsx_rows(content)
    else:
        raise ValueError("仅支持 .xls/.xlsx 往来科目明细文件")

    parent_code = ""
    parent_name = ""
    warnings: list[str] = []
    parsed_rows: list[LegacySubjectRow] = []
    seen_codes: set[str] = set()
    first_parent_code = ""
    first_parent_name = ""

    for row_index, row in enumerate(rows, start=1):
        code = row[0] if row else ""
        name = row[1] if len(row) > 1 else ""
        if not code and not name:
            continue

        split = _split_code_name(code)
        if split and not name:
            code, name = split
        if _looks_like_header(code, name):
            continue

        if (split and len(code) == 4) or (re.fullmatch(r"\d{4}", code or "") and name):
            parent_code = code
            parent_name = name
            first_parent_code = first_parent_code or parent_code
            first_parent_name = first_parent_name or parent_name
            continue

        if not parent_code:
            warnings.append(f"第 {row_index} 行在识别到父级科目前出现，已跳过")
            continue
        if not code or not name:
            warnings.append(f"第 {row_index} 行缺少编码或名称，已跳过")
            continue
        if not re.fullmatch(r"[0-9A-Za-z.\-_]+", code):
            warnings.append(f"第 {row_index} 行编码格式无法识别：{code}，已跳过")
            continue

        full_code, row_parent_code, row_parent_name = _resolve_child_code(parent_code, parent_name, code)
        if full_code in seen_codes:
            warnings.append(f"第 {row_index} 行编码 {full_code} 重复，已跳过")
            continue
        seen_codes.add(full_code)
        parsed_rows.append(
            LegacySubjectRow(
                code=code,
                name=name,
                full_code=full_code,
                parent_code=row_parent_code,
                parent_name=row_parent_name,
            )
        )

    if not first_parent_code or not first_parent_name:
        raise ValueError("未识别到父级往来科目，请确认文件包含 1122/1221/2202/2241 等父级科目行")
    if not parsed_rows:
        raise ValueError("未识别到往来明细科目")

    return LegacySubjectWorkbook(
        parent_code=first_parent_code,
        parent_name=first_parent_name,
        rows=parsed_rows,
        warnings=warnings,
    )


async def _find_subject_in_scope(
    db: AsyncSession, code: str, client_id: str | None
) -> AccountSubject | None:
    result = await db.execute(
        select(AccountSubject).where(AccountSubject.code == code, _scope_clause(client_id))
    )
    return result.scalar_one_or_none()


async def _subject_used_by_entries(db: AsyncSession, code: str) -> bool:
    result = await db.execute(
        select(func.count()).select_from(JournalEntryLine).where(
            JournalEntryLine.account_code == code
        )
    )
    return bool(result.scalar() or 0)


async def _ensure_parent(
    db: AsyncSession,
    *,
    parent_code: str,
    parent_name: str,
    client_id: str | None,
) -> None:
    category, direction = _category_direction(parent_code)
    existing = await _find_subject_in_scope(db, parent_code, client_id)
    if not existing and client_id:
        existing = await _find_subject_in_scope(db, parent_code, None)
    if existing:
        existing.full_name = existing.full_name or parent_name
        existing.is_leaf = False
        existing.is_active = True
        return

    db.add(
        AccountSubject(
            id=str(uuid.uuid4()),
            client_id=client_id,
            code=parent_code,
            name=parent_name,
            full_name=parent_name,
            level=1,
            parent_code=None,
            parent_account_name=None,
            category=category,
            direction=direction,
            is_leaf=False,
            is_active=True,
            created_from="legacy_import",
        )
    )


async def import_legacy_subjects(
    db: AsyncSession,
    *,
    filename: str,
    content: bytes,
    client_id: str | None = None,
) -> SubjectImportResponse:
    workbook = parse_legacy_subject_workbook(filename, content)
    response = SubjectImportResponse(
        filename=filename,
        parent_code=workbook.parent_code,
        parent_name=workbook.parent_name,
        warnings=list(workbook.warnings),
    )

    ensured_parents: set[str] = set()
    for row in workbook.rows:
        category, direction = _category_direction(row.parent_code)
        if row.parent_code not in ensured_parents:
            await _ensure_parent(
                db,
                parent_code=row.parent_code,
                parent_name=row.parent_name,
                client_id=client_id,
            )
            ensured_parents.add(row.parent_code)

        full_name = f"{row.parent_name}_{row.name}"
        existing = await _find_subject_in_scope(db, row.full_code, client_id)
        if existing:
            if normalize_party_name(existing.name) != normalize_party_name(row.name):
                used = await _subject_used_by_entries(db, row.full_code)
                reason = "编码已被凭证使用，不能覆盖名称" if used else "同一公司下编码已存在但名称不同"
                response.conflicts.append(
                    SubjectImportConflict(
                        code=row.full_code,
                        name=row.name,
                        reason=reason,
                        existing_code=existing.code,
                        existing_name=existing.name,
                    )
                )
                continue
            existing.full_name = full_name
            existing.parent_code = row.parent_code
            existing.parent_account_name = row.parent_name
            existing.category = category
            existing.direction = direction
            existing.level = 2
            existing.is_leaf = True
            existing.is_active = True
            existing.created_from = existing.created_from or "legacy_import"
            response.updated += 1
            continue

        name_result = await db.execute(
            select(AccountSubject).where(
                _scope_clause(client_id),
                AccountSubject.parent_code == row.parent_code,
                AccountSubject.name == row.name,
                AccountSubject.code != row.full_code,
                AccountSubject.is_active == True,
            )
        )
        same_name = name_result.scalar_one_or_none()
        if same_name:
            response.skipped += 1
            response.conflicts.append(
                SubjectImportConflict(
                    code=row.full_code,
                    name=row.name,
                    reason="同一父级下名称已存在但编码不同，未自动覆盖",
                    existing_code=same_name.code,
                    existing_name=same_name.name,
                )
            )
            continue

        db.add(
            AccountSubject(
                id=str(uuid.uuid4()),
                client_id=client_id,
                code=row.full_code,
                name=row.name,
                full_name=full_name,
                level=2,
                parent_code=row.parent_code,
                parent_account_name=row.parent_name,
                category=category,
                direction=direction,
                is_leaf=True,
                is_active=True,
                created_from="legacy_import",
            )
        )
        response.inserted += 1

    await db.flush()
    return response


async def find_legacy_sub_account_by_counterparty(
    db: AsyncSession,
    *,
    client_id: str,
    counterparty_name: str | None,
    allowed_parent_codes: list[str] | tuple[str, ...] | None = None,
) -> AccountSubject | None:
    party = (counterparty_name or "").strip()
    if not party:
        return None
    allowed = tuple(allowed_parent_codes or CURRENT_PARENT_CODES)
    if not allowed:
        return None

    result = await db.execute(
        select(AccountSubject)
        .where(
            AccountSubject.is_active == True,
            AccountSubject.is_leaf == True,
            AccountSubject.level == 2,
            AccountSubject.parent_code.in_(allowed),
            (AccountSubject.client_id == client_id) | (AccountSubject.client_id == None),
        )
        .order_by(AccountSubject.client_id.desc(), AccountSubject.code)
    )
    candidates = list(result.scalars().all())

    for subject in candidates:
        if (subject.name or "").strip() == party:
            return subject

    compact_party = normalize_party_name(party)
    for subject in candidates:
        if normalize_party_name(subject.name) == compact_party:
            return subject

    containing = [
        subject for subject in candidates
        if compact_party and compact_party in normalize_party_name(subject.full_name)
    ]
    return containing[0] if len(containing) == 1 else None


def legacy_subject_line_fields(
    subject: AccountSubject,
    *,
    is_income: bool,
) -> dict[str, str | None]:
    parent_code = subject.parent_code or subject.code[:4]
    parent_name = subject.parent_account_name or CURRENT_PARENT_NAMES.get(parent_code)
    full_name = subject.full_name or (
        f"{parent_name}_{subject.name}" if parent_name else subject.name
    )
    auxiliary_type = "customer" if parent_code.startswith(("1122", "1123", "1221")) else "supplier"
    return {
        "account_code": subject.code,
        "account_name": subject.name,
        "account_full_name": full_name,
        "parent_account_code": parent_code,
        "parent_account_name": parent_name,
        "auxiliary_type": auxiliary_type,
        "auxiliary_code": subject.code,
        "auxiliary_name": subject.name,
    }
