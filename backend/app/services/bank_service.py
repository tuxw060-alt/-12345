"""
Bank statement import and auto-matching service.

Parses bank CSV/Excel files (ICBC, CCB, ABC, etc.) and suggests
journal entries based on transaction descriptions.
"""

import csv
import io
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession
from loguru import logger

# Transaction category → account code mapping
TRANSACTION_RULES = [
    # 收入类
    {"kw": "货款|销售款|服务费|咨询费|收入|汇款|转入|收到", "code": "5001", "name": "主营业务收入", "dir": "credit"},
    {"kw": "利息收入|利息|结息|存款利息", "code": "5603.01", "name": "财务费用-利息收入", "dir": "credit"},
    {"kw": "退税|税收返还|财政补贴|政府补助", "code": "5301", "name": "营业外收入", "dir": "credit"},
    # 支出类
    {"kw": "工资|奖金|津贴|薪酬", "code": "2211.01", "name": "应付职工薪酬-工资", "dir": "debit"},
    {"kw": "社保|社会保险|养老|医疗|失业|工伤|生育", "code": "2211.02", "name": "应付职工薪酬-社保", "dir": "debit"},
    {"kw": "公积金|住房公积金", "code": "2211.03", "name": "应付职工薪酬-公积金", "dir": "debit"},
    {"kw": "税金|缴税|纳税|扣税|增值税|所得税|印花税|附加税|城建税|教育费", "code": "2221", "name": "应交税费", "dir": "debit"},
    {"kw": "房租|租金|租赁|物业费|水电费|电费|水费|燃气", "code": "5602", "name": "管理费用", "dir": "debit"},
    {"kw": "办公|文具|打印|墨盒|硒鼓|快递|邮寄|顺丰|中通|圆通", "code": "5602.02", "name": "管理费用-办公费", "dir": "debit"},
    {"kw": "差旅|住宿|机票|火车票|高铁|酒店|宾馆|打车|滴滴", "code": "5602.04", "name": "管理费用-差旅费", "dir": "debit"},
    {"kw": "餐费|餐饮|招待|宴请|饭店|酒楼|餐厅", "code": "5602.05", "name": "管理费用-业务招待费", "dir": "debit"},
    {"kw": "手续费|账户管理费|网银费|回单柜|对公账户", "code": "5603.02", "name": "财务费用-手续费", "dir": "debit"},
    {"kw": "还款|贷款|借款|还本|付息", "code": "2001", "name": "短期借款", "dir": "debit"},
    {"kw": "充值|提现|转账|转存", "code": "1002", "name": "银行存款", "dir": "debit"},
    {"kw": "通讯|电话费|话费|宽带|网费|电信|移动|联通", "code": "5602.16", "name": "管理费用-通讯费", "dir": "debit"},
    {"kw": "维修|修理|保养|维护|配件|零件", "code": "5602.14", "name": "管理费用-维修费", "dir": "debit"},
]


def parse_bank_csv(file_content: bytes, filename: str) -> list[dict[str, Any]]:
    """Parse a bank CSV file into transaction records."""
    text = file_content.decode("gbk", errors="ignore") or file_content.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    if len(rows) < 2:
        return []

    # Detect header row — find which columns contain 日期/摘要/收入/支出/余额
    header = rows[0]
    col_map: dict[str, int | None] = {"date": None, "desc": None, "income": None, "expense": None, "balance": None, "counterparty": None}

    for i, h in enumerate(header):
        h_lower = h.strip().lower().replace(" ", "")
        if any(k in h_lower for k in ["日期", "交易时间", "记账日期", "date"]): col_map["date"] = i
        elif any(k in h_lower for k in ["摘要", "用途", "说明", "交易说明", "desc", "memo"]): col_map["desc"] = i
        elif any(k in h_lower for k in ["收入", "贷方", "存入", "credit", "in"]): col_map["income"] = i
        elif any(k in h_lower for k in ["支出", "借方", "取出", "debit", "out"]): col_map["expense"] = i
        elif any(k in h_lower for k in ["余额", "balance"]): col_map["balance"] = i
        elif any(k in h_lower for k in ["对方", "户名", "名称", "counterparty"]): col_map["counterparty"] = i

    # If no header found, try treating first row as data with fixed positions
    if not col_map["date"] and not col_map["desc"]:
        # Try common 银行 fixed format: date, desc, income, expense, balance
        if len(header) >= 4:
            col_map = {"date": 0, "desc": 1, "income": 2, "expense": 3, "balance": 4, "counterparty": None}

    transactions = []
    for row in rows[1:]:
        if not row or len(row) < 3:
            continue

        def get_val(idx: int | None) -> str:
            if idx is None or idx >= len(row):
                return ""
            return row[idx].strip().replace('"', "").replace("'", "")

        desc = get_val(col_map.get("desc"))
        if not desc or "合计" in desc or "小计" in desc:
            continue

        # Parse amounts
        income_str = get_val(col_map.get("income")).replace(",", "").replace("¥", "")
        expense_str = get_val(col_map.get("expense")).replace(",", "").replace("¥", "")

        try:
            income = float(income_str) if income_str else 0
        except ValueError:
            income = 0
        try:
            expense = float(expense_str) if expense_str else 0
        except ValueError:
            expense = 0

        if income == 0 and expense == 0:
            continue

        # Parse date
        date_str = get_val(col_map.get("date"))
        parsed_date = _parse_date(date_str)

        transactions.append({
            "date": parsed_date,
            "description": desc,
            "income": income,
            "expense": expense,
            "balance": get_val(col_map.get("balance")),
            "counterparty": get_val(col_map.get("counterparty")),
        })

    return transactions


def parse_bank_excel(file_content: bytes, filename: str) -> list[dict[str, Any]]:
    """Parse a bank Excel file into transaction records."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active

    rows = [[str(c.value or "") for c in r] for r in ws.iter_rows()]
    if len(rows) < 2:
        return []

    # Same header detection as CSV
    col_map = {"date": 0, "desc": 1, "income": 2, "expense": 3, "balance": 4, "counterparty": None}
    header = rows[0]
    for i, h in enumerate(header):
        h_lower = h.strip().lower().replace(" ", "")
        if any(k in h_lower for k in ["日期", "交易时间"]): col_map["date"] = i
        elif any(k in h_lower for k in ["摘要", "用途", "说明"]): col_map["desc"] = i
        elif any(k in h_lower for k in ["收入", "贷方", "存入"]): col_map["income"] = i
        elif any(k in h_lower for k in ["支出", "借方", "取出"]): col_map["expense"] = i
        elif any(k in h_lower for k in ["余额"]): col_map["balance"] = i
        elif any(k in h_lower for k in ["对方", "户名"]): col_map["counterparty"] = i

    transactions = []
    for row in rows[1:]:
        if not row:
            continue

        def gv(i): return row[i].strip() if i < len(row) else ""

        desc = gv(col_map["desc"])
        if not desc or "合计" in desc or "小计" in desc:
            continue

        try:
            income = float(gv(col_map["income"]).replace(",", "").replace("¥", "")) if gv(col_map["income"]) else 0
        except ValueError:
            income = 0
        try:
            expense = float(gv(col_map["expense"]).replace(",", "").replace("¥", "")) if gv(col_map["expense"]) else 0
        except ValueError:
            expense = 0

        if income == 0 and expense == 0:
            continue

        transactions.append({
            "date": _parse_date(gv(col_map["date"])),
            "description": desc,
            "income": income,
            "expense": expense,
            "balance": gv(col_map["balance"]),
            "counterparty": gv(col_map["counterparty"]),
        })

    return transactions


def _parse_date(date_str: str) -> str:
    """Try to parse various date formats to YYYY-MM-DD."""
    import re
    # YYYY-MM-DD, YYYY/MM/DD, YYYYMMDD, MM/DD/YYYY, DD/MM/YYYY
    patterns = [
        (r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", lambda m: f"{m[0]}-{m[1]:0>2}-{m[2]:0>2}"),
        (r"(\d{1,2})/(\d{1,2})/(\d{4})", lambda m: f"{m[2]}-{m[0]:0>2}-{m[1]:0>2}"),
        (r"(\d{4})(\d{2})(\d{2})", lambda m: f"{m[0]}-{m[1]}-{m[2]}"),
    ]
    for pat, fmt in patterns:
        match = re.search(pat, date_str)
        if match:
            try:
                g = match.groups()
                return fmt(g)
            except (ValueError, IndexError):
                pass
    return date.today().isoformat()


def match_transactions(transactions: list[dict]) -> list[dict]:
    """Auto-match bank transactions to account subjects."""
    for txn in transactions:
        desc = txn["description"]
        matched = None

        for rule in TRANSACTION_RULES:
            import re
            if re.search(rule["kw"], desc):
                matched = rule
                break

        if matched:
            txn["suggested_code"] = matched["code"]
            txn["suggested_name"] = matched["name"]
            txn["suggested_dir"] = matched["dir"]
            txn["suggested_amount"] = txn["income"] if matched["dir"] == "credit" else txn["expense"]
            txn["auto_matched"] = True
        else:
            # Default: income → 主营业务收入, expense → 管理费用-其他
            if txn["income"] > 0:
                txn["suggested_code"] = "5001"
                txn["suggested_name"] = "主营业务收入"
                txn["suggested_dir"] = "credit"
                txn["suggested_amount"] = txn["income"]
            else:
                txn["suggested_code"] = "5602.99"
                txn["suggested_name"] = "管理费用-其他"
                txn["suggested_dir"] = "debit"
                txn["suggested_amount"] = txn["expense"]
            txn["auto_matched"] = False

    return transactions
