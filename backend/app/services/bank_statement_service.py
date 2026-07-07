"""Bank statement upload, AI extraction, and entry generation."""

import csv
import io
import re
import uuid
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import settings
from app.models.account_subject import AccountSubject
from app.models.bank_statement import BankStatementTransaction, BankStatementUpload
from app.models.journal_entry import JournalEntry
from app.schemas.journal_entry import EntryCreate, EntryLineCreate
from app.services.ai_service import ai_service, extract_text
from app.services.entry_service import create_entry
from app.services.document_voucher_service import (
    identify_business_type,
    recommend_template,
    generate_voucher_draft_from_document,
)
from app.schemas.document_voucher import TemplatePreviewRequest
from app.services.subject_service import (
    CURRENT_PARENT_CODES as LEGACY_CURRENT_PARENT_CODES,
    find_legacy_sub_account_by_counterparty,
    legacy_subject_line_fields,
)


SPREADSHEET_EXTENSIONS = {".csv", ".xlsx", ".xls", ".xlsm", ".ods"}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}
BANK_ACCOUNT_CODE = "100201"
BANK_ACCOUNT_NAME = "银行存款_基本户"
BANK_ACCOUNT_AUX = "基本户"
UNRECOGNIZED_COUNTERPARTY = "对方未识别"
RECEIVABLE_PAYABLE_PREFIXES = ("1122", "2202", "1123", "2203", "1221", "2241")
CURRENT_PARENT_CODES = {"1122", "1123", "1221", "2202", "2203", "2241"}


def detect_bank_statement_processing_mode(filename: str | Path | None) -> dict[str, Any]:
    ext = Path(str(filename or "")).suffix.lower()
    if ext in {".xls", ".xlsx", ".xlsm", ".ods"}:
        return {
            "file_type": "excel",
            "processing_mode": "excel_parser",
            "use_ocr": False,
            "use_ai": False,
            "display_text": "正在解析 Excel 流水",
            "description": "Excel 流水优先使用表格解析，不走 OCR。",
        }
    if ext == ".csv":
        return {
            "file_type": "csv",
            "processing_mode": "csv_parser",
            "use_ocr": False,
            "use_ai": False,
            "display_text": "正在解析 CSV 流水",
            "description": "CSV 流水优先使用表格解析，不走 OCR。",
        }
    if ext == ".pdf":
        return {
            "file_type": "pdf",
            "processing_mode": "pdf_text_parser",
            "use_ocr": False,
            "use_ai": False,
            "display_text": "正在解析 PDF 文本流水",
            "description": "优先抽取 PDF 文本和表格，失败后再 OCR。",
        }
    if ext in IMAGE_EXTENSIONS:
        return {
            "file_type": "image",
            "processing_mode": "image_ocr",
            "use_ocr": True,
            "use_ai": True,
            "display_text": "正在 OCR 识别图片流水",
            "description": "图片流水使用 OCR + AI 字段整理。",
        }
    return {
        "file_type": "unknown",
        "processing_mode": "unsupported",
        "use_ocr": False,
        "use_ai": False,
        "display_text": "不支持的文件类型",
        "description": "请上传 xls、xlsx、csv、pdf、jpg、jpeg、png、webp 文件。",
    }


def _set_upload_processing(upload: BankStatementUpload, mode: dict[str, Any]) -> None:
    upload.file_type = mode["file_type"]
    upload.processing_mode = mode["processing_mode"]
    upload.use_ocr = bool(mode["use_ocr"])
    upload.use_ai = bool(mode["use_ai"])
    upload.processing_display = mode["display_text"]
    upload.processing_description = mode["description"]


def _is_bank_fee(summary: str | None) -> bool:
    """Check if a transaction summary indicates a bank fee.

    NOTE: "服务费" is intentionally excluded — it's too broad and would
    falsely match non-fee service transactions. Only bank-specific fee
    keywords are matched.
    """
    text = (summary or "").lower()
    return any(keyword in text for keyword in (
        "手续费", "手續費", "银行收费", "对公收费", "對公收費",
        "账户管理费", "賬戶管理費", "短信费", "网银服务费", "網銀服務費",
        "汇款费", "转账费", "电子汇划费", "電子匯劃費",
        "回单", "对公账户", "对公維護", "扣费", "扣費",
        "560301", "5603.02",
    ))


def _raw_value(tx: BankStatementTransaction, key: str) -> Any:
    return (tx.raw_data or {}).get(key)


def _tx_counterparty(tx: BankStatementTransaction) -> str:
    party = _clean_party_name(tx.counterparty)
    if party:
        return party
    if _is_bank_fee(tx.summary):
        bank_name = _raw_value(tx, "bank_name")
        return str(bank_name).strip() if bank_name else "银行手续费"
    return UNRECOGNIZED_COUNTERPARTY


def _account_full_name(name: str | None) -> str:
    return (name or "").replace("-", "_")


def _is_receivable_payable_account(account_code: str | None) -> bool:
    return any((account_code or "").startswith(prefix) for prefix in RECEIVABLE_PAYABLE_PREFIXES)


def normalize_account_name(name: str | None, auxiliary_name: str | None, account_code: str | None) -> str:
    account_name = _account_full_name(name).strip()
    aux_name = (auxiliary_name or "").strip()
    if not account_name:
        return ""
    if not _is_receivable_payable_account(account_code) or not aux_name:
        return account_name
    underscore_suffix = f"_{aux_name}"
    if account_name.endswith(underscore_suffix):
        return account_name[: -len(underscore_suffix)].rstrip("_锛?-")
    if account_name.endswith(aux_name):
        return account_name[: -len(aux_name)].rstrip("_锛?-")
    return account_name


def _aux_type_for_account(account_code: str, is_income: bool) -> str:
    if account_code.startswith(("1122", "1123", "1221")):
        return "customer"
    if account_code.startswith(("2202", "2203", "2241")):
        return "supplier"
    if account_code.startswith("100"):
        return "bank_account"
    return "counterparty" if not is_income else "customer"


def _compact_party_name(value: str | None) -> str:
    return re.sub(r"\s+", "", value or "")


def _parent_code_for_current_account(account_code: str | None, is_income: bool) -> str:
    code = account_code or ""
    for prefix in RECEIVABLE_PAYABLE_PREFIXES:
        if code.startswith(prefix):
            return prefix
    return "1122" if is_income else "2202"


def _allowed_parent_codes_for_transaction(
    tx: BankStatementTransaction,
    line_account_code: str | None,
) -> list[str]:
    summary = (tx.summary or "").replace(" ", "")
    suggested = _parent_code_for_current_account(line_account_code, bool(tx.income_amount))
    if any(word in summary for word in ("还款", "往来款", "借款", "归还借款", "代垫", "备用金")):
        preferred = ["2241", "1221", "2202", "1122", "2203", "1123"]
    elif tx.income_amount:
        preferred = ["1122", "1221", "2203", "1123", "2241", "2202"]
    else:
        preferred = ["2202", "2241", "1221", "1122", "2203", "1123"]
    ordered = [suggested, *preferred, *LEGACY_CURRENT_PARENT_CODES]
    return list(dict.fromkeys(code for code in ordered if code))


async def _match_current_account_subject(
    db: AsyncSession,
    *,
    client_id: str,
    parent_code: str,
    counterparty_name: str,
) -> AccountSubject | None:
    party = (counterparty_name or "").strip()
    if not party or party == UNRECOGNIZED_COUNTERPARTY:
        return None

    stmt = (
        select(AccountSubject)
        .where(
            AccountSubject.is_active == True,
            AccountSubject.is_leaf == True,
            AccountSubject.parent_code == parent_code,
            (AccountSubject.client_id == client_id) | (AccountSubject.client_id == None),
        )
        .order_by(AccountSubject.client_id.desc(), AccountSubject.code)
    )
    result = await db.execute(stmt)
    candidates = list(result.scalars().all())
    for subject in candidates:
        if (subject.name or "").strip() == party:
            return subject

    compact_party = _compact_party_name(party)
    for subject in candidates:
        if _compact_party_name(subject.name) == compact_party:
            return subject
    return None


def _subject_line_from_match(
    subject: AccountSubject,
    fallback_parent_name: str | None,
    is_income: bool,
) -> dict[str, str | None]:
    parent_name = subject.parent_account_name or fallback_parent_name
    full_name = subject.full_name or (
        f"{parent_name}_{subject.name}" if parent_name else subject.name
    )
    return {
        "account_code": subject.code,
        "account_name": subject.name,
        "account_full_name": full_name,
        "parent_account_code": subject.parent_code,
        "parent_account_name": parent_name,
        "auxiliary_type": _aux_type_for_account(subject.code, is_income=is_income),
        "auxiliary_code": subject.code,
        "auxiliary_name": subject.name,
    }


def _apply_account_fields_to_line(line: EntryLineCreate, fields: dict[str, str | None]) -> None:
    line.account_code = str(fields["account_code"] or line.account_code)
    line.account_name = str(fields["account_name"] or line.account_name)
    line.account_full_name = fields.get("account_full_name")
    line.parent_account_code = fields.get("parent_account_code")
    line.parent_account_name = fields.get("parent_account_name")
    line.auxiliary_type = fields.get("auxiliary_type")
    line.auxiliary_code = fields.get("auxiliary_code")
    line.auxiliary_name = fields.get("auxiliary_name")


def _selected_account_fields(tx: BankStatementTransaction, is_income: bool) -> dict[str, str | None] | None:
    if not tx.manual_account_override or not tx.selected_account_code:
        return None
    parent_code = tx.selected_parent_account_code
    if not parent_code:
        for prefix in RECEIVABLE_PAYABLE_PREFIXES:
            if tx.selected_account_code.startswith(prefix):
                parent_code = prefix
                break
    auxiliary_type = _aux_type_for_account(tx.selected_account_code, is_income=is_income)
    return {
        "account_code": tx.selected_account_code,
        "account_name": tx.selected_account_name or tx.selected_account_full_name or tx.selected_account_code,
        "account_full_name": tx.selected_account_full_name or tx.selected_account_name,
        "parent_account_code": parent_code,
        "parent_account_name": tx.selected_parent_account_name,
        "auxiliary_type": auxiliary_type,
        "auxiliary_code": tx.selected_account_code if parent_code else None,
        "auxiliary_name": tx.selected_account_name if parent_code else None,
    }


def map_bank_statement_columns(headers: list[Any]) -> dict[str, Any]:
    normalized = [str(item or "").strip().lower().replace(" ", "") for item in headers]
    aliases: dict[str, tuple[str, ...]] = {
        "transactionDate": ("日期", "交易日期", "记账日期", "入账日期", "交易时间", "date"),
        "summary": ("摘要", "交易摘要", "用途", "交易用途", "备注", "附言", "desc", "memo"),
        "counterpartyName": ("对方户名", "对方名称", "对方账户名称", "对方单位", "交易对方", "counterparty"),
        "counterpartyAccount": ("对方账号", "对方账户", "账号", "account"),
        "incomeAmount": ("收入", "收入金额", "贷方发生额", "收方金额", "入账金额", "贷方"),
        "expenseAmount": ("支出", "支出金额", "借方发生额", "付方金额", "出账金额", "借方"),
        "balance": ("余额", "账户余额", "当前余额", "balance"),
        "serialNo": ("流水号", "交易流水号", "交易编号", "凭证号", "回单编号"),
        "purpose": ("用途", "交易用途", "附言", "备注"),
    }
    result: dict[str, Any] = {key: None for key in aliases}
    ambiguous: list[str] = []
    for key, words in aliases.items():
        matches = [
            headers[index]
            for index, header in enumerate(normalized)
            if header and any(word.lower() in header for word in words)
        ]
        if len(matches) == 1:
            result[key] = matches[0]
        elif len(matches) > 1:
            result[key] = matches[0]
            ambiguous.append(key)
    required = ("transactionDate", "summary", "incomeAmount", "expenseAmount")
    missing = [key for key in required if not result.get(key)]
    result["confidence"] = max(0, round((len(aliases) - len(missing) - len(ambiguous) * 0.5) / len(aliases) * 100, 2))
    result["missingFields"] = missing
    result["ambiguousFields"] = ambiguous
    return result


def _to_float(value: Any) -> float | None:
    """Parse monetary value from cell, handling mixed formats.

    Handles: "1,234.56", "¥12.00", " -8.00 ", "￥100.50", "1,000",
             "(100.00)", negative signs, whitespace, and currency symbols.
    """
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not _is_finite_float(value):
            return None
        return float(value)
    if _looks_like_time(value):
        return None
    text = str(value).strip()
    # Detect parenthesized negative: "(100.00)" -> -100.00
    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]
    # Remove currency symbols and spacing
    text = re.sub(r"[¥￥$€£\s]", "", text)
    # Remove thousands separators but keep decimal point and minus sign
    text = text.replace(",", "")
    # Detect trailing minus or leading minus
    if text.endswith("-"):
        negative = True
        text = text[:-1]
    if text.startswith("-"):
        negative = True
        text = text[1:]
    # Keep only digits and decimal point
    text = re.sub(r"[^\d.]", "", text)
    if not text or text == ".":
        return None
    try:
        result = float(text)
    except ValueError:
        return None
    if negative:
        result = -result
    if not _is_finite_float(result):
        return None
    return result


def _is_finite_float(value: float) -> bool:
    import math
    return math.isfinite(value)


def _looks_like_time(value: Any) -> bool:
    return bool(re.fullmatch(r"\s*\d{1,2}:\d{2}(:\d{2})?\s*", str(value or "")))


def _to_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    text = str(value).strip().replace("/", "-").replace(".", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y%m%d"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            continue
    return None


def _rows_to_text(rows: list[list[Any]]) -> str:
    return "\n".join("\t".join("" if c is None else str(c) for c in row) for row in rows)


def _clean_party_name(counterparty: str | None) -> str:
    party = (counterparty or "").strip()
    if not party:
        return ""
    if re.fullmatch(r"[\d,.\-+\s:]+", party):
        return ""
    bank_noise = ("银行", "银联", "财付通", "支付宝", "微信支付")
    if any(word in party for word in bank_noise) and len(party) <= 12:
        return ""
    return party


def _extract_ods_rows(path: Path) -> list[list[Any]]:
    namespaces = {
        "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
        "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
        "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
    }
    table_ns = f"{{{namespaces['table']}}}"
    office_ns = f"{{{namespaces['office']}}}"
    text_ns = f"{{{namespaces['text']}}}"
    rows: list[list[Any]] = []

    with zipfile.ZipFile(path) as ods:
        with ods.open("content.xml") as content:
            root = ET.parse(content).getroot()

    for table in root.findall(".//table:table", namespaces):
        table_name = table.attrib.get(f"{table_ns}name")
        if table_name:
            rows.append([f"宸ヤ綔琛?{table_name}"])
        for row in table.findall("table:table-row", namespaces):
            repeat_rows = min(int(row.attrib.get(f"{table_ns}number-rows-repeated", "1")), 500)
            values: list[Any] = []
            for cell in row.findall("table:table-cell", namespaces):
                repeat_cols = min(int(cell.attrib.get(f"{table_ns}number-columns-repeated", "1")), 50)
                value: Any = cell.attrib.get(f"{office_ns}value")
                if value is None:
                    value = cell.attrib.get(f"{office_ns}date-value")
                if value is None:
                    parts = [
                        "".join(p.itertext())
                        for p in cell.findall(f"{text_ns}p")
                    ]
                    value = "\n".join(part for part in parts if part).strip() or None
                values.extend([value] * repeat_cols)
            if any(v not in (None, "") for v in values):
                for _ in range(repeat_rows):
                    rows.append(values)
            if len(rows) >= 500:
                break
        if len(rows) >= 500:
            break

    return rows[:500]


def _guess_subject(
    summary: str | None,
    counterparty: str | None,
    is_income: bool,
) -> tuple[str, str, str]:
    text = f"{summary or ""} {counterparty or ""}".lower()
    if _is_bank_fee(text):
        return "560301", "财务费用-手续费", "手续费"
    if any(keyword in text for keyword in ("还款", "往来款", "借款", "归还", "代垫", "备用金")):
        party = _clean_party_name(counterparty)
        return "2241", f"其他应付款_{party}" if party else "其他应付款", "往来款"
    if any(keyword in text for keyword in ("贷款", "银行借款", "借款本金")):
        return "2001", "短期借款", "银行贷款"
    if is_income:
        if any(keyword in text for keyword in ("利息", "结息")):
            return "560303", "财务费用-利息收入", "利息收入"
        party = _clean_party_name(counterparty)
        return "1122", f"应收账款_{party}" if party else "应收账款", "银行收款"
    rules = [
        (("利息",), "560301", "财务费用-利息支出", "利息支出"),
        (("工资", "薪酬", "代发"), "221101", "应付职工薪酬-工资", "工资薪酬"),
        (("社保", "养老", "医保"), "221102", "应付职工薪酬-社保", "社保款项"),
        (("公积金",), "221103", "应付职工薪酬-公积金", "公积金款项"),
        (("税", "税费", "国库"), "2221", "应交税费", "税费缴纳"),
        (("办公", "文具", "打印", "耗材"), "560202", "管理费用-办公费", "办公支出"),
        (("餐", "饮", "招待", "宴请"), "560205", "管理费用-业务招待费", "餐饮招待"),
        (("差旅", "住宿", "酒店", "机票", "火车", "打车"), "560204", "管理费用-差旅费", "差旅出行"),
        (("加油", "停车", "etc", "过路"), "560203", "管理费用-交通费", "交通车辆"),
        (("租金", "房租", "租赁"), "560208", "管理费用-租赁费", "租赁支出"),
        (("软件", "平台", "技术服务", "saas"), "560211", "管理费用-软件服务费", "软件技术服务"),
        (("咨询", "审计", "律师", "代理记账", "知识产权"), "560210", "管理费用-中介咨询费", "专业服务"),
    ]
    for keywords, code, name, reason in rules:
        if any(keyword in text for keyword in keywords):
            return code, name, reason
    party = _clean_party_name(counterparty)
    return "2202", f"应付账款_{party}" if party else "应付账款", "银行付款"

def _detect_columns(rows: list[list[Any]]) -> dict[str, int | None]:
    """Detect column positions by scanning first 10 rows for headers and typed values."""
    col_map: dict[str, int | None] = {
        "date": None, "desc": None, "income": None, "expense": None,
        "balance": None, "counterparty": None, "account": None,
    }
    # First, try header matching
    header_keywords = {
        "date": ("日期", "交易时间", "记账日期", "date", "时间"),
        "desc": ("摘要", "用途", "说明", "交易说明", "desc", "memo", "摘要/用途", "交易摘要"),
        "income": ("收入", "贷方", "存入", "credit", "收", "收入金额", "贷方金额"),
        "expense": ("支出", "借方", "取出", "debit", "付", "支出金额", "借方金额"),
        "balance": ("余额", "balance", "账户余额"),
        "counterparty": ("对方", "户名", "名称", "counterparty", "对方户名", "对方名称", "交易对方"),
        "account": ("账号", "对方账号", "卡号", "account"),
    }
    # Scan first 10 rows for headers
    for row in rows[:10]:
        if not row:
            continue
        for i, cell in enumerate(row):
            if cell is None:
                continue
            text = str(cell).strip().lower().replace(" ", "").replace("\n", "")
            for key, keywords in header_keywords.items():
                if col_map[key] is not None:
                    continue
                if any(kw in text for kw in keywords):
                    col_map[key] = i
        # If all found, break
        if all(v is not None for v in col_map.values()):
            break

    for row in rows[:80]:
        if (
            len(row) >= 8
            and _to_date(row[0])
            and _looks_like_time(row[1])
            and (_to_float(row[2]) is not None or _to_float(row[3]) is not None)
            and _to_float(row[4]) is not None
        ):
            col_map.update({
                "date": 0,
                "expense": 2,
                "income": 3,
                "balance": 4,
                "account": 5,
                "counterparty": 6,
                "desc": 7,
            })
            break

    # If headers not found, try typed-value detection
    if col_map["date"] is None or col_map["desc"] is None:
        for row in rows[:50]:
            if not row or len(row) < 4:
                continue
            if len(row) >= 8 and _to_date(row[0]) and _looks_like_time(row[1]):
                col_map.update({
                    "date": 0,
                    "expense": 2,
                    "income": 3,
                    "balance": 4,
                    "account": 5,
                    "counterparty": 6,
                    "desc": 7,
                })
                break
            for i, cell in enumerate(row):
                if cell is None:
                    continue
                if col_map["date"] is None and _to_date(cell):
                    col_map["date"] = i
                elif col_map["expense"] is None and _to_float(cell) and _to_float(cell) > 0:
                    # Check next column for income
                    col_map["expense"] = i
                elif col_map["income"] is None and _to_float(cell) and col_map["expense"] is not None and i != col_map["expense"]:
                    col_map["income"] = i
            if col_map["date"] is not None:
                break

    # Fallback: common Chinese bank format: date, desc, expense, income, balance, ..., counterparty
    if col_map["date"] is None:
        col_map["date"] = 0
    if col_map["desc"] is None:
        col_map["desc"] = 1
    if col_map["expense"] is None:
        col_map["expense"] = 2
    if col_map["income"] is None:
        col_map["income"] = 3
    if col_map["balance"] is None:
        col_map["balance"] = 4

    return col_map


def _parse_transaction_rows(rows: list[list[Any]]) -> list[dict[str, Any]]:
    """Parse bank export rows with smart column detection."""
    col_map = _detect_columns(rows)
    transactions: list[dict[str, Any]] = []

    for row in rows:
        if len(row) < 3:
            continue

        def gv(idx: int | None) -> Any:
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        tx_date = _to_date(gv(col_map["date"]))
        if not tx_date:
            continue

        expense = _to_float(gv(col_map["expense"])) if col_map["expense"] is not None else None
        income = _to_float(gv(col_map["income"])) if col_map["income"] is not None else None

        # Sometimes income/expense are in same column (signed)
        if not income and not expense:
            val = _to_float(gv(col_map.get("expense") or col_map.get("income")))
            if val is not None:
                expense = abs(val) if val < 0 else None
                income = val if val > 0 else None
        if not income and not expense:
            continue

        summary = str(gv(col_map["desc"])).strip() if col_map["desc"] is not None and gv(col_map["desc"]) not in (None, "") else None
        counterparty = str(gv(col_map["counterparty"])).strip() if col_map["counterparty"] is not None and gv(col_map["counterparty"]) not in (None, "") else None
        balance = _to_float(gv(col_map["balance"]))
        account_number = str(gv(col_map["account"])).strip() if col_map["account"] is not None and gv(col_map["account"]) not in (None, "") else None

        code, name, reason = _guess_subject(summary, counterparty, bool(income))

        transactions.append({
            "transaction_date": tx_date.isoformat(),
            "summary": summary,
            "counterparty": counterparty,
            "account_number": account_number,
            "income_amount": income,
            "expense_amount": expense,
            "balance": balance,
            "suggested_subject_code": code,
            "suggested_subject_name": name,
            "subject_reason": reason,
            "confidence": 100,
            "source": "spreadsheet",
            "raw_row": row,
        })
    return transactions


def _spreadsheet_rows(path: Path) -> tuple[str, list[list[Any]]]:
    if path.suffix.lower() == ".csv":
        raw = path.read_bytes()
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw.decode("utf-8", errors="ignore")
        rows = list(csv.reader(io.StringIO(text)))
        return text, rows

    if path.suffix.lower() == ".ods":
        rows = _extract_ods_rows(path)
        return _rows_to_text(rows), rows

    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[list[Any]] = []
    try:
        for ws in wb.worksheets:
            rows.append([f"宸ヤ綔琛?{ws.title}"])
            for row in ws.iter_rows(max_row=500, values_only=True):
                if any(cell not in (None, "") for cell in row):
                    rows.append(list(row))
    finally:
        wb.close()
    return _rows_to_text(rows), rows


def parse_bank_statement_csv(path: Path) -> tuple[str, list[dict[str, Any]], int]:
    text, rows = _spreadsheet_rows(path)
    return text, _parse_transaction_rows(rows), len(rows)


def parse_bank_statement_excel(path: Path) -> tuple[str, list[dict[str, Any]], int]:
    text, rows = _spreadsheet_rows(path)
    return text, _parse_transaction_rows(rows), len(rows)


def _extract_spreadsheet(path: Path) -> tuple[str, list[dict[str, Any]]]:
    if path.suffix.lower() == ".csv":
        raw = path.read_bytes()
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw.decode("utf-8", errors="ignore")
        rows = list(csv.reader(io.StringIO(text)))
        return _rows_to_text(rows[:500]), _parse_transaction_rows(rows)

    if path.suffix.lower() == ".ods":
        rows = _extract_ods_rows(path)
        return _rows_to_text(rows), _parse_transaction_rows(rows)

    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[list[Any]] = []
    try:
        for ws in wb.worksheets:
            rows.append([f"宸ヤ綔琛? {ws.title}"])
            for row in ws.iter_rows(max_row=500, values_only=True):
                if any(cell not in (None, "") for cell in row):
                    rows.append(list(row))
    finally:
        wb.close()
    return _rows_to_text(rows), _parse_transaction_rows(rows)


def _guess_subject(
    summary: str | None,
    counterparty: str | None,
    is_income: bool,
) -> tuple[str, str, str]:
    text = f"{summary or ""} {counterparty or ""}".lower()
    if _is_bank_fee(text):
        return "560301", "财务费用-手续费", "手续费"
    if any(keyword in text for keyword in ("还款", "往来款", "借款", "归还", "代垫", "备用金")):
        party = _clean_party_name(counterparty)
        return "2241", f"其他应付款_{party}" if party else "其他应付款", "往来款"
    if any(keyword in text for keyword in ("贷款", "银行借款", "借款本金")):
        return "2001", "短期借款", "银行贷款"
    if is_income:
        if any(keyword in text for keyword in ("利息", "结息")):
            return "560303", "财务费用-利息收入", "利息收入"
        party = _clean_party_name(counterparty)
        return "1122", f"应收账款_{party}" if party else "应收账款", "银行收款"
    rules = [
        (("利息",), "560301", "财务费用-利息支出", "利息支出"),
        (("工资", "薪酬", "代发"), "221101", "应付职工薪酬-工资", "工资薪酬"),
        (("社保", "养老", "医保"), "221102", "应付职工薪酬-社保", "社保款项"),
        (("公积金",), "221103", "应付职工薪酬-公积金", "公积金款项"),
        (("税", "税费", "国库"), "2221", "应交税费", "税费缴纳"),
        (("办公", "文具", "打印", "耗材"), "560202", "管理费用-办公费", "办公支出"),
        (("餐", "饮", "招待", "宴请"), "560205", "管理费用-业务招待费", "餐饮招待"),
        (("差旅", "住宿", "酒店", "机票", "火车", "打车"), "560204", "管理费用-差旅费", "差旅出行"),
        (("加油", "停车", "etc", "过路"), "560203", "管理费用-交通费", "交通车辆"),
        (("租金", "房租", "租赁"), "560208", "管理费用-租赁费", "租赁支出"),
        (("软件", "平台", "技术服务", "saas"), "560211", "管理费用-软件服务费", "软件技术服务"),
        (("咨询", "审计", "律师", "代理记账", "知识产权"), "560210", "管理费用-中介咨询费", "专业服务"),
    ]
    for keywords, code, name, reason in rules:
        if any(keyword in text for keyword in keywords):
            return code, name, reason
    party = _clean_party_name(counterparty)
    return "2202", f"应付账款_{party}" if party else "应付账款", "银行付款"

def _detect_columns(rows: list[list[Any]]) -> dict[str, int | None]:
    """Detect bank statement columns with readable Chinese/English headers."""
    col_map: dict[str, int | None] = {
        "date": None,
        "desc": None,
        "income": None,
        "expense": None,
        "balance": None,
        "counterparty": None,
        "account": None,
    }
    header_keywords = {
        "date": ("日期", "交易时间", "交易日期", "记账日期", "入账日期", "date", "时间"),
        "desc": ("摘要", "用途", "说明", "交易说明", "备注", "附言", "desc", "memo", "交易摘要"),
        "income": ("收入", "贷方", "存入", "收方", "收款", "转入", "credit", "收入金额", "贷方金额"),
        "expense": ("支出", "借方", "取出", "付方", "付款", "转出", "debit", "支出金额", "借方金额"),
        "balance": ("余额", "账户余额", "当前余额", "balance"),
        "counterparty": ("对方", "户名", "名称", "counterparty", "对方户名", "对方名称", "交易对方", "收款人", "付款人"),
        "account": ("账号", "账户", "对方账号", "对方账户", "卡号", "account"),
    }

    for row in rows[:10]:
        for index, cell in enumerate(row):
            text = str(cell or "").strip().lower().replace(" ", "").replace("\n", "")
            if not text:
                continue
            for key, keywords in header_keywords.items():
                if col_map[key] is None and any(keyword in text for keyword in keywords):
                    col_map[key] = index
        if col_map["date"] is not None and (col_map["income"] is not None or col_map["expense"] is not None):
            break

    for row in rows[:80]:
        if (
            len(row) >= 8
            and _to_date(row[0])
            and _looks_like_time(row[1])
            and (_to_float(row[2]) is not None or _to_float(row[3]) is not None)
            and _to_float(row[4]) is not None
        ):
            col_map.update({
                "date": 0,
                "expense": 2,
                "income": 3,
                "balance": 4,
                "account": 5,
                "counterparty": 6,
                "desc": 7,
            })
            return col_map

    if col_map["date"] is None or (col_map["income"] is None and col_map["expense"] is None):
        for row in rows[:50]:
            if not row or len(row) < 4:
                continue
            for index, cell in enumerate(row):
                if col_map["date"] is None and _to_date(cell):
                    col_map["date"] = index
                elif _to_float(cell) is not None:
                    if col_map["expense"] is None:
                        col_map["expense"] = index
                    elif col_map["income"] is None and index != col_map["expense"]:
                        col_map["income"] = index
            if col_map["date"] is not None and (col_map["income"] is not None or col_map["expense"] is not None):
                break

    col_map["date"] = col_map["date"] if col_map["date"] is not None else 0
    col_map["desc"] = col_map["desc"] if col_map["desc"] is not None else 1
    col_map["expense"] = col_map["expense"] if col_map["expense"] is not None else 2
    col_map["income"] = col_map["income"] if col_map["income"] is not None else 3
    col_map["balance"] = col_map["balance"] if col_map["balance"] is not None else 4
    return col_map


def _spreadsheet_rows(path: Path) -> tuple[str, list[list[Any]]]:
    """Read CSV, ODS, legacy XLS, and modern Excel rows."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        raw = path.read_bytes()
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw.decode("utf-8", errors="ignore")
        rows = list(csv.reader(io.StringIO(text)))
        return text, rows

    if suffix == ".ods":
        rows = _extract_ods_rows(path)
        return _rows_to_text(rows), rows

    if suffix == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(str(path))
        rows: list[list[Any]] = []
        for sheet in workbook.sheets():
            rows.append([f"宸ヤ綔琛?{sheet.name}"])
            for row_index in range(min(sheet.nrows, 500)):
                row = sheet.row_values(row_index)
                if any(cell not in (None, "") for cell in row):
                    rows.append(row)
        return _rows_to_text(rows), rows

    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[list[Any]] = []
    try:
        for ws in wb.worksheets:
            rows.append([f"宸ヤ綔琛?{ws.title}"])
            for row in ws.iter_rows(max_row=500, values_only=True):
                if any(cell not in (None, "") for cell in row):
                    rows.append(list(row))
    finally:
        wb.close()
    return _rows_to_text(rows), rows


def _group_transactions_for_entries(
    transactions: list[BankStatementTransaction],
) -> list[tuple[EntryCreate, list[BankStatementTransaction]]]:
    """Group transactions by subject code and return entry payloads with source rows."""
    from collections import defaultdict

    groups: dict[tuple[str, str, str], list[BankStatementTransaction]] = defaultdict(list)
    for tx in transactions:
        if tx.entry_id or tx.status != "recognized":
            continue
        code = tx.suggested_subject_code or ("1122" if tx.income_amount else "2202")
        name = tx.suggested_subject_name or ("搴旀敹璐︽" if tx.income_amount else "搴斾粯璐︽")
        direction = "expense" if tx.expense_amount else "income"
        if direction == "income" and code.startswith("1122"):
            key = ("__bank_receipt__", "银行收款", direction)
        elif direction == "expense" and code.startswith("2202"):
            key = ("__bank_payment__", "银行付款", direction)
        else:
            key = (code, name, direction)
        groups[key].append(tx)

    grouped_entries: list[tuple[EntryCreate, list[BankStatementTransaction]]] = []
    for (code, name, direction), txs in groups.items():
        total = sum(float(t.expense_amount or t.income_amount or 0) for t in txs)
        if total <= 0:
            continue

        # Build merged summary: 银行X鏈堟墜缁垂锛圢绗旓級
        tx_dates = [_to_date(t.transaction_date) for t in txs if t.transaction_date]
        tx_dates = [d for d in tx_dates if d]
        first_date = min(tx_dates) if tx_dates else date.today()
        last_date = max(tx_dates) if tx_dates else date.today()
        month_str = f"{first_date.month}月" if first_date.month == last_date.month else f"{first_date.month}-{last_date.month}月"
        bank_code, bank_name = BANK_ACCOUNT_CODE, BANK_ACCOUNT_NAME

        if code == "__bank_receipt__":
            summary = "银行收款"
            counterparty_totals: dict[tuple[str, str], float] = defaultdict(float)
            for tx in txs:
                line_code = tx.suggested_subject_code or "1122"
                line_name = tx.suggested_subject_name or "搴旀敹璐︽"
                counterparty_totals[(line_code, line_name)] += float(tx.income_amount or 0)
            lines = [
                EntryLineCreate(line_number=1, account_code=bank_code, account_name=bank_name,
                                direction="debit", debitAmount=round(total, 2), creditAmount=0,
                                summary_detail=summary),
            ]
            for index, ((line_code, line_name), amount) in enumerate(counterparty_totals.items(), start=2):
                lines.append(EntryLineCreate(line_number=index, account_code=line_code, account_name=line_name,
                                             direction="credit", debitAmount=0, creditAmount=round(amount, 2),
                                             summary_detail=summary))
        elif code == "__bank_payment__":
            summary = "银行付款"
            counterparty_totals: dict[tuple[str, str], float] = defaultdict(float)
            for tx in txs:
                line_code = tx.suggested_subject_code or "2202"
                line_name = tx.suggested_subject_name or "搴斾粯璐︽"
                counterparty_totals[(line_code, line_name)] += float(tx.expense_amount or 0)
            lines = [
                EntryLineCreate(line_number=index, account_code=line_code, account_name=line_name,
                                direction="debit", debitAmount=round(amount, 2), creditAmount=0,
                                summary_detail=summary)
                for index, ((line_code, line_name), amount) in enumerate(counterparty_totals.items(), start=1)
            ]
            lines.append(EntryLineCreate(line_number=len(lines) + 1, account_code=bank_code, account_name=bank_name,
                                         direction="credit", debitAmount=0, creditAmount=round(total, 2),
                                         summary_detail=summary))
        elif direction == "expense":
            summary = "手续费" if code.startswith(("5603.02", "560301")) else f"{name}（{month_str}，{len(txs)}笔）"
            lines = [
                EntryLineCreate(line_number=1, account_code=code, account_name=name,
                                direction="debit", debitAmount=round(total, 2), creditAmount=0,
                                summary_detail=summary),
                EntryLineCreate(line_number=2, account_code=bank_code, account_name=bank_name,
                                direction="credit", debitAmount=0, creditAmount=round(total, 2),
                                summary_detail=summary),
            ]
        else:
            summary = "结息" if code.startswith("5603.03") else f"{name}（{month_str}，{len(txs)}笔）"
            lines = [
                EntryLineCreate(line_number=1, account_code=bank_code, account_name=bank_name,
                                direction="debit", debitAmount=round(total, 2), creditAmount=0,
                                summary_detail=summary),
                EntryLineCreate(line_number=2, account_code=code, account_name=name,
                                direction="credit", debitAmount=0, creditAmount=round(total, 2),
                                summary_detail=summary),
            ]

        grouped_entries.append((
            EntryCreate(
                client_id=txs[0].client_id,
                voucher_date=last_date or date.today(),
                voucher_type="记",
                summary=summary[:500],
                lines=lines,
            ),
            txs,
        ))

    return grouped_entries


def _transaction_to_entry(
    tx: BankStatementTransaction,
    voucher_type: str = "记",
) -> EntryCreate | None:
    amount = float(tx.expense_amount or tx.income_amount or 0)
    if amount <= 0:
        return None

    counterparty_name = _tx_counterparty(tx)
    counterparty_account = (tx.account_number or "").strip() or None
    is_income = bool(tx.income_amount)

    if is_income:
        subject_code = tx.suggested_subject_code or "1122"
        subject_name = normalize_account_name(tx.suggested_subject_name or "搴旀敹璐︽", counterparty_name, subject_code)
        summary = tx.summary or "银行收款"
    else:
        subject_code = tx.suggested_subject_code or "2202"
        subject_name = normalize_account_name(tx.suggested_subject_name or "搴斾粯璐︽", counterparty_name, subject_code)
        summary = tx.summary or "银行付款"
        if _is_bank_fee(tx.summary) or subject_code.startswith(("5603.02", "560301")):
            subject_code = "560301"
            subject_name = "财务费用-手续费"
            summary = tx.summary or "手续费"

    subject_aux_type = _aux_type_for_account(subject_code, is_income=is_income)
    subject_aux_name = counterparty_name if _is_receivable_payable_account(subject_code) else None
    if counterparty_name == UNRECOGNIZED_COUNTERPARTY:
        subject_aux_name = counterparty_name if subject_code.startswith(("1122", "2202", "1123", "2203")) else None

    source_fields = {
        "counterparty_name": counterparty_name,
        "counterparty_account": counterparty_account,
        "source_type": "bank_statement",
        "source_document_id": tx.upload_id,
        "source_row_id": tx.id,
    }
    bank_line_fields = {
        "account_code": BANK_ACCOUNT_CODE,
        "account_name": BANK_ACCOUNT_NAME,
        "account_full_name": BANK_ACCOUNT_NAME,
        "auxiliary_type": "bank_account",
        "auxiliary_name": BANK_ACCOUNT_AUX,
        **source_fields,
    }
    subject_line_fields = {
        "account_code": subject_code,
        "account_name": subject_name,
        "account_full_name": subject_name,
        "auxiliary_type": subject_aux_type,
        "auxiliary_name": subject_aux_name,
        **source_fields,
    }

    if tx.expense_amount:
        lines = [
            EntryLineCreate(
                line_number=1,
                direction="debit",
                debitAmount=round(amount, 2),
                creditAmount=0,
                summary_detail=summary,
                **subject_line_fields,
            ),
            EntryLineCreate(
                line_number=2,
                direction="credit",
                debitAmount=0,
                creditAmount=round(amount, 2),
                summary_detail=summary,
                **bank_line_fields,
            ),
        ]
    else:
        lines = [
            EntryLineCreate(
                line_number=1,
                direction="debit",
                debitAmount=round(amount, 2),
                creditAmount=0,
                summary_detail=summary,
                **bank_line_fields,
            ),
            EntryLineCreate(
                line_number=2,
                direction="credit",
                debitAmount=0,
                creditAmount=round(amount, 2),
                summary_detail=summary,
                **subject_line_fields,
            ),
        ]

    return EntryCreate(
        client_id=tx.client_id,
        voucher_date=tx.transaction_date or date.today(),
        voucher_type=voucher_type,
        summary=summary[:500],
        lines=lines,
    )


async def _transaction_to_entry_with_subject(
    db: AsyncSession,
    tx: BankStatementTransaction,
    voucher_type: str = "记",
) -> EntryCreate | None:
    amount = float(tx.expense_amount or tx.income_amount or 0)
    if amount <= 0:
        return None

    selected_fields = _selected_account_fields(tx, is_income=bool(tx.income_amount))
    document = {
        "template_id": tx.selected_template_id or tx.recommended_template_id,
        "document_type_id": tx.document_type_id,
        "document_name": tx.document_name or "银行票据",
        "settlement_method": tx.settlement_method or "银行",
        "business_type": tx.business_type,
        "summary": tx.summary,
        "counterparty_name": _tx_counterparty(tx),
        "counterparty_account": tx.account_number,
        "income_amount": float(tx.income_amount or 0),
        "expense_amount": float(tx.expense_amount or 0),
        "balance": float(tx.balance or 0) if tx.balance is not None else None,
        "source_type": "bank_statement",
        "source_document_id": tx.upload_id,
        "source_row_id": tx.id,
        "selected_account": selected_fields,
    }
    return await generate_voucher_draft_from_document(
        db,
        client_id=tx.client_id,
        document=document,
        voucher_date=tx.transaction_date or date.today(),
        voucher_type=voucher_type,
    )


async def upload_and_extract(
    db: AsyncSession,
    file: UploadFile,
    client_id: str,
    auto_generate: bool = False,
    merge_similar: bool = True,
) -> tuple[BankStatementUpload, list[str]]:
    upload_id = str(uuid.uuid4())
    original_filename = file.filename or "bank-statement.xlsx"
    ext = Path(original_filename).suffix.lower() or ".xlsx"
    save_name = f"{upload_id}{ext}"
    upload_path = settings.upload_path / save_name
    content = await file.read()
    upload_path.write_bytes(content)
    processing_mode = detect_bank_statement_processing_mode(original_filename)

    upload = BankStatementUpload(
        id=upload_id,
        client_id=client_id,
        file_path=str(upload_path),
        filename=original_filename,
        status="pending",
    )
    _set_upload_processing(upload, processing_mode)
    db.add(upload)
    await db.flush()

    try:
        local_transactions: list[dict[str, Any]] = []
        total_rows = 0
        if upload.processing_mode == "csv_parser":
            text, local_transactions, total_rows = parse_bank_statement_csv(upload_path)
        elif upload.processing_mode == "excel_parser":
            text, local_transactions, total_rows = parse_bank_statement_excel(upload_path)
        else:
            text = extract_text(upload_path)
        upload.raw_text = text[:60000]
        upload.total_rows = total_rows or None
        upload.valid_rows = len(local_transactions) if local_transactions else 0
        upload.error_rows = max(0, (total_rows or 0) - (upload.valid_rows or 0)) if total_rows else None

        if upload.processing_mode in {"excel_parser", "csv_parser"}:
            result = {
                "warning": None,
                "transactions": local_transactions,
                "processing_mode": upload.processing_mode,
                "use_ai": False,
            }
        elif settings.deepseek_api_key:
            result = await ai_service.extract_bank_statement(text)
        else:
            result = {
                "warning": "未配置DEEPSEEK_API_KEY，已使用本地规则解析流水，AI科目推荐未启用",
                "transactions": local_transactions,
            }
        upload.raw_ai_response = result
        if "error" in result and not local_transactions:
            upload.status = "failed"
            upload.error_msg = result["error"]
            await db.flush()
            return await get_upload(db, upload.id), []
        if "error" in result and local_transactions:
            result = {
                "warning": result["error"],
                "transactions": local_transactions,
            }
            upload.raw_ai_response = result

        transactions = result.get("transactions") or local_transactions
        if not transactions:
            upload.status = "failed"
            upload.error_msg = "鏈瘑鍒埌银行娴佹按鏄庣粏"
            await db.flush()
            return await get_upload(db, upload.id), []

        for item in transactions:
            income = _to_float(item.get("income_amount"))
            expense = _to_float(item.get("expense_amount"))
            income = abs(income) if income else None
            expense = abs(expense) if expense else None
            if income and expense:
                if abs(income) >= abs(expense):
                    expense = None
                else:
                    income = None

            tx = BankStatementTransaction(
                id=str(uuid.uuid4()),
                upload_id=upload.id,
                client_id=client_id,
                transaction_date=_to_date(item.get("transaction_date")),
                summary=item.get("summary"),
                counterparty=item.get("counterparty") or item.get("counterpartyName"),
                account_number=item.get("account_number") or item.get("counterpartyAccount"),
                income_amount=income,
                expense_amount=expense,
                balance=_to_float(item.get("balance")),
                suggested_subject_code=item.get("suggested_subject_code"),
                suggested_subject_name=item.get("suggested_subject_name"),
                subject_reason=item.get("subject_reason"),
                confidence=_to_float(item.get("confidence")),
                status="recognized" if (income or expense) else "failed",
                error_msg=None if (income or expense) else "鏈瘑鍒埌浜ゆ槗閲戦",
                raw_data=item,
            )
            if income or expense:
                business_type = identify_business_type(
                    " ".join(str(v or "") for v in (tx.summary, tx.counterparty)),
                    is_income=bool(income),
                )
                recommendation = await recommend_template(
                    db,
                    TemplatePreviewRequest(
                        client_id=client_id,
                        document_name="银行票据",
                        settlement_method="银行",
                        business_type=business_type,
                        summary=tx.summary,
                        counterparty_name=tx.counterparty,
                        income_amount=income,
                        expense_amount=expense,
                        balance=float(tx.balance or 0) if tx.balance is not None else None,
                    ),
                )
                tx.document_type_id = recommendation.document_type_id
                tx.document_name = recommendation.document_name or "银行票据"
                tx.settlement_method = recommendation.settlement_method or "银行"
                tx.business_type = recommendation.business_type
                tx.recommended_template_id = recommendation.template_id
                tx.selected_template_id = recommendation.template_id
                tx.template_match_reason = recommendation.reason
            db.add(tx)

        upload.status = "done"
        await db.flush()
    except Exception as e:
        upload.status = "failed"
        upload.error_msg = str(e)
        await db.flush()
        return await get_upload(db, upload.id), []

    entry_ids: list[str] = []
    upload = await get_upload(db, upload.id)
    if auto_generate:
        consumed_ids: set[str] = set()
        if merge_similar:
            merged_ids, consumed_ids = await _create_merged_fee_entries(db, list(upload.transactions))
            entry_ids.extend(merged_ids)
        for tx in upload.transactions:
            if tx.id in consumed_ids:
                continue
            if tx.status == "recognized" and not tx.entry_id:
                entry_data = await _transaction_to_entry_with_subject(db, tx)
                if entry_data:
                    entry = await create_entry(db, entry_data)
                    tx.entry_id = entry.id
                    tx.status = "voucher_created"
                    entry_ids.append(entry.id)
        await db.flush()
        upload = await get_upload(db, upload.id)

    return upload, entry_ids


async def list_uploads(
    db: AsyncSession,
    client_id: str | None = None,
    status: str | None = None,
    offset: int = 0,
    limit: int = 50,
):
    stmt = select(BankStatementUpload).options(
        selectinload(BankStatementUpload.transactions)
    )
    if client_id:
        stmt = stmt.where(BankStatementUpload.client_id == client_id)
    if status:
        stmt = stmt.where(BankStatementUpload.status == status)

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await db.execute(count_stmt)).scalar() or 0
    stmt = stmt.order_by(BankStatementUpload.created_at.desc()).offset(offset).limit(limit)
    result = await db.execute(stmt)
    return result.unique().scalars().all(), total


async def get_upload(db: AsyncSession, upload_id: str) -> BankStatementUpload | None:
    stmt = (
        select(BankStatementUpload)
        .options(selectinload(BankStatementUpload.transactions))
        .where(BankStatementUpload.id == upload_id)
    )
    result = await db.execute(stmt)
    return result.unique().scalar_one_or_none()


async def delete_upload(db: AsyncSession, upload: BankStatementUpload) -> None:
    await db.delete(upload)
    await db.flush()


async def get_transaction(
    db: AsyncSession, transaction_id: str
) -> BankStatementTransaction | None:
    stmt = select(BankStatementTransaction).where(
        BankStatementTransaction.id == transaction_id
    )
    result = await db.execute(stmt)
    return result.scalar_one_or_none()


async def update_transaction_account_selection(
    db: AsyncSession,
    tx: BankStatementTransaction,
    *,
    account_code: str,
    account_name: str,
    account_full_name: str | None = None,
    parent_account_code: str | None = None,
    parent_account_name: str | None = None,
    source: str = "manual",
) -> BankStatementTransaction:
    tx.selected_account_code = account_code
    tx.selected_account_name = account_name
    tx.selected_account_full_name = account_full_name or account_name
    tx.selected_parent_account_code = parent_account_code
    tx.selected_parent_account_name = parent_account_name
    tx.manual_account_override = True
    tx.account_selection_source = source
    tx.suggested_subject_code = account_code
    tx.suggested_subject_name = account_full_name or account_name
    await db.flush()
    return tx


async def update_transaction_template_selection(
    db: AsyncSession,
    tx: BankStatementTransaction,
    *,
    document_type_id: str | None = None,
    document_name: str | None = None,
    settlement_method: str | None = None,
    business_type: str | None = None,
    template_id: str | None = None,
) -> BankStatementTransaction:
    tx.document_type_id = document_type_id or tx.document_type_id
    tx.document_name = document_name or tx.document_name
    tx.settlement_method = settlement_method or tx.settlement_method
    tx.business_type = business_type or tx.business_type
    tx.selected_template_id = template_id
    if not template_id:
        recommendation = await recommend_template(
            db,
            TemplatePreviewRequest(
                client_id=tx.client_id,
                document_type_id=tx.document_type_id,
                document_name=tx.document_name,
                settlement_method=tx.settlement_method,
                business_type=tx.business_type,
                summary=tx.summary,
                counterparty_name=tx.counterparty,
                income_amount=float(tx.income_amount or 0),
                expense_amount=float(tx.expense_amount or 0),
                balance=float(tx.balance or 0) if tx.balance is not None else None,
            ),
        )
        tx.recommended_template_id = recommendation.template_id
        tx.selected_template_id = recommendation.template_id
        tx.template_match_reason = recommendation.reason
    else:
        tx.template_match_reason = "浜哄伐閫夋嫨妯℃澘"
    await db.flush()
    return tx


async def generate_entry_for_transaction(
    db: AsyncSession, tx: BankStatementTransaction
) -> str:
    if tx.entry_id == "merged":
        tx.entry_id = None
    if tx.entry_id:
        existing = await db.get(JournalEntry, tx.entry_id)
        if existing:
            return tx.entry_id
        tx.entry_id = None
    entry_data = await _transaction_to_entry_with_subject(db, tx)
    if not entry_data:
        raise ValueError("该流水无法生成凭证")
    entry = await create_entry(db, entry_data)
    tx.entry_id = entry.id
    tx.status = "voucher_created"
    await db.flush()
    return entry.id


def _fee_merge_key(tx: BankStatementTransaction) -> tuple | None:
    """Return a merge key for fee transactions.

    Requires a reliable serial number to merge. Without it, each fee row stays
    independent — identical amounts alone do NOT justify merging.
    """
    if tx.entry_id or tx.status != "recognized":
        return None
    if not tx.transaction_date or not tx.expense_amount:
        return None
    if tx.manual_account_override and (tx.selected_account_code or tx.suggested_subject_code) not in {"560301", "5603.02"}:
        return None
    business_type = (tx.business_type or "").strip()
    subject_code = (tx.selected_account_code or tx.suggested_subject_code or "").strip()
    if business_type != "手续费" and not _is_bank_fee(tx.summary) and subject_code not in {"560301", "5603.02"}:
        return None
    if subject_code and subject_code not in {"560301", "5603.02"}:
        return None
    # Require serial number to merge — without it, don't merge.
    serial = (_raw_value(tx, "serial_no") or _raw_value(tx, "流水号") or "").strip()
    if not serial:
        return None
    counterparty = (tx.counterparty or "").strip()
    return (
        tx.client_id,
        tx.transaction_date.isoformat(),
        counterparty,
        serial,
        "手续费",
        "560301",
        "100201",
        "CNY",
    )


async def _create_merged_fee_entries(
    db: AsyncSession,
    transactions: list[BankStatementTransaction],
) -> tuple[list[str], set[str]]:
    from collections import defaultdict

    groups: dict[tuple, list[BankStatementTransaction]] = defaultdict(list)
    for tx in transactions:
        key = _fee_merge_key(tx)
        if key:
            groups[key].append(tx)

    entry_ids: list[str] = []
    consumed_ids: set[str] = set()
    for txs in groups.values():
        if len(txs) < 2:
            continue
        total = round(sum(float(tx.expense_amount or 0) for tx in txs), 2)
        if total <= 0:
            continue
        source_row_ids = [tx.id for tx in txs]
        n_tx = len(txs)
        month_str = f"{txs[0].transaction_date.month}月" if txs[0].transaction_date else ""
        entry_data = EntryCreate(
            client_id=txs[0].client_id,
            voucher_date=txs[0].transaction_date or date.today(),
            voucher_type="记",
            summary=f"手续费（{month_str}，{n_tx}笔）" if month_str else f"手续费（{n_tx}笔）",
            sourceRowIds=source_row_ids,
            lines=[
                EntryLineCreate(
                    line_number=1,
                    account_code="560301",
                    account_name="财务费用-手续费",
                    account_full_name="财务费用-手续费",
                    direction="debit",
                    debitAmount=total,
                    creditAmount=0,
                    summary_detail=f"手续费（{n_tx}笔）",
                    source_type="bank_statement",
                    source_document_id=txs[0].upload_id,
                ),
                EntryLineCreate(
                    line_number=2,
                    account_code=BANK_ACCOUNT_CODE,
                    account_name=BANK_ACCOUNT_NAME,
                    account_full_name=BANK_ACCOUNT_NAME,
                    auxiliary_type="bank_account",
                    auxiliary_name=BANK_ACCOUNT_AUX,
                    direction="credit",
                    debitAmount=0,
                    creditAmount=total,
                    summary_detail=f"手续费（{n_tx}笔）",
                    source_type="bank_statement",
                    source_document_id=txs[0].upload_id,
                ),
            ],
        )
        entry = await create_entry(db, entry_data)
        for tx in txs:
            tx.entry_id = entry.id
            tx.status = "voucher_created"
            consumed_ids.add(tx.id)
        entry_ids.append(entry.id)
    return entry_ids, consumed_ids


async def generate_entries_for_client(
    db: AsyncSession,
    client_id: str,
    merge_similar: bool = True,
) -> list[str]:
    stmt = (
        select(BankStatementTransaction)
        .where(
            BankStatementTransaction.client_id == client_id,
            BankStatementTransaction.status == "recognized",
        )
        .order_by(BankStatementTransaction.transaction_date, BankStatementTransaction.created_at)
    )
    result = await db.execute(stmt)
    transactions = list(result.scalars().all())

    existing_ids = {tx.entry_id for tx in transactions if tx.entry_id and tx.entry_id != "merged"}
    valid_ids: set[str] = set()
    if existing_ids:
        existing_result = await db.execute(
            select(JournalEntry.id).where(JournalEntry.id.in_(existing_ids))
        )
        valid_ids = set(existing_result.scalars().all())

    for tx in transactions:
        if tx.entry_id == "merged" or (tx.entry_id and tx.entry_id not in valid_ids):
            tx.entry_id = None

    entry_ids: list[str] = []
    consumed_ids: set[str] = set()
    if merge_similar:
        merged_ids, consumed_ids = await _create_merged_fee_entries(db, transactions)
        entry_ids.extend(merged_ids)

    for tx in transactions:
        if tx.id in consumed_ids:
            continue
        if tx.status == "recognized" and not tx.entry_id:
            entry_data = await _transaction_to_entry_with_subject(db, tx)
            if entry_data:
                entry = await create_entry(db, entry_data)
                tx.entry_id = entry.id
                tx.status = "voucher_created"
                entry_ids.append(entry.id)

    await db.flush()
    return entry_ids
