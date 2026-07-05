"""Kingdee 快记帐 voucher-import Excel export service."""

import json
import re
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Sequence

from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from loguru import logger

from app.config import settings
from app.models.journal_entry import JournalEntry, JournalEntryLine


# 快记帐「凭证导入模板.xlsx」列顺序。会计要求客户/供应商列填辅助核算编码。
KINGDEE_HEADERS = [
    "日期",
    "凭证字",
    "凭证号",
    "录入顺序",
    "摘要",
    "科目编码",
    "科目名称",
    "借方金额",
    "贷方金额",
    "客户",
    "供应商",
]

CUSTOMER_ACCOUNT_PREFIXES = ("1122", "1123", "1221")
SUPPLIER_ACCOUNT_PREFIXES = ("2202", "2203", "2241")
BANK_ACCOUNT_PREFIXES = ("1001", "1002", "100201")
GENERIC_COUNTERPARTY_VALUES = {"", "支付", "收付", "银行流水", "销售", "付款", "收款"}
AUX_REGISTRY_FILE = "kingdee_auxiliary_registry.json"
TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent.parent
    / "assets"
    / "templates"
    / "kingdee_voucher_import_template.xlsx"
)
DATA_START_ROW = 2


def _registry_path() -> Path:
    return settings.data_path / AUX_REGISTRY_FILE


def _load_aux_registry() -> dict:
    path = _registry_path()
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        logger.warning(f"Auxiliary registry is unreadable, rebuilding: {path}")
        return {}


def _save_aux_registry(registry: dict) -> None:
    path = _registry_path()
    path.write_text(
        json.dumps(registry, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def _normalize_aux_name(name: str) -> str:
    return re.sub(r"\s+", "", name.strip())


def _clean_counterparty(value: str | None) -> str:
    if not value:
        return ""
    return " ".join(str(value).replace("\n", " ").replace("\r", " ").split())[:200]


def _entry_counterparty(entry: JournalEntry) -> str:
    for line in entry.lines:
        counterparty = _clean_counterparty(getattr(line, "counterparty_name", None))
        if counterparty and counterparty not in GENERIC_COUNTERPARTY_VALUES:
            return counterparty

    for line in entry.lines:
        if not line.account_code.startswith(BANK_ACCOUNT_PREFIXES):
            continue
        counterparty = _clean_counterparty(line.summary_detail)
        if counterparty and counterparty not in GENERIC_COUNTERPARTY_VALUES:
            return counterparty

    summary = _clean_counterparty(entry.summary)
    parenthesized = re.search(r"\(([^()]+)\)\s*$", summary)
    if parenthesized:
        return _clean_counterparty(parenthesized.group(1))
    if " - " in summary:
        return _clean_counterparty(summary.rsplit(" - ", 1)[1])
    return ""


def _extract_aux_name(line: JournalEntryLine) -> str | None:
    """Return the auxiliary name embedded in account name, e.g. 应收账款_某公司."""
    if getattr(line, "auxiliary_name", None):
        return _clean_counterparty(line.auxiliary_name)
    for sep in ("_", "＿"):
        if sep in line.account_name:
            name = line.account_name.rsplit(sep, 1)[-1].strip()
            return name or None
    return None


def _aux_category(account_code: str) -> str | None:
    if account_code.startswith(CUSTOMER_ACCOUNT_PREFIXES):
        return "客户"
    if account_code.startswith(SUPPLIER_ACCOUNT_PREFIXES):
        return "供应商"
    return None


def _next_aux_code(items: dict[str, str]) -> str:
    used_numbers = []
    for code in items.values():
        if isinstance(code, str) and code.isdigit():
            used_numbers.append(int(code))
    return f"{(max(used_numbers) if used_numbers else 0) + 1:03d}"


def _get_aux_code(registry: dict, client_id: str, category: str, name: str) -> str:
    client_registry = registry.setdefault(client_id or "default", {})
    category_registry = client_registry.setdefault(category, {})
    normalized_name = _normalize_aux_name(name)
    if normalized_name not in category_registry:
        category_registry[normalized_name] = _next_aux_code(category_registry)
    return category_registry[normalized_name]


def _load_export_workbook():
    if TEMPLATE_PATH.exists():
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
        if ws.max_row > DATA_START_ROW:
            ws.delete_rows(DATA_START_ROW + 1, ws.max_row - DATA_START_ROW)
        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            for cell in ws[row_idx]:
                cell.value = None
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "凭证导入模板"
    return wb, ws


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, len(KINGDEE_HEADERS) + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
        if source.comment:
            target.comment = copy(source.comment)


async def get_entries_for_export(
    db: AsyncSession,
    client_id: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    status: str = "confirmed",
) -> list[JournalEntry]:
    """Fetch confirmed entries ready for export."""
    stmt = (
        select(JournalEntry)
        .where(JournalEntry.status == status)
        .order_by(JournalEntry.voucher_date, JournalEntry.created_at)
    )
    if client_id:
        stmt = stmt.where(JournalEntry.client_id == client_id)
    if date_from:
        stmt = stmt.where(JournalEntry.voucher_date >= date_from)
    if date_to:
        stmt = stmt.where(JournalEntry.voucher_date <= date_to)

    result = await db.execute(stmt)
    entries = result.scalars().all()

    # Eager-load lines
    for entry in entries:
        # Access lines to ensure they're loaded
        _ = entry.lines

    return list(entries)


def generate_kingdee_excel(
    entries: Sequence[JournalEntry],
    output_path: str | Path | None = None,
) -> Path:
    """
    Generate a Kingdee-compatible Excel file from journal entries.

    Args:
        entries: Confirmed journal entries with lines.
        output_path: Where to save the file. Auto-generates if None.

    Returns:
        Path to the generated .xlsx file.
    """
    wb, ws = _load_export_workbook()

    aux_registry = _load_aux_registry()
    registry_changed = False

    # === Styles for fallback workbook without template ===
    header_font = Font(name="微软雅黑", bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="center")
    amount_alignment = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # === Write header row ===
    for col_idx, header in enumerate(KINGDEE_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        if not cell.value:
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    # === Write data rows ===
    row = DATA_START_ROW
    for entry in entries:
        voucher_date_str = entry.voucher_date.isoformat() if entry.voucher_date else ""
        voucher_number = entry.voucher_number or ""
        entry_counterparty = _entry_counterparty(entry)

        for line_index, line in enumerate(entry.lines, 1):
            if row != DATA_START_ROW:
                _copy_row_style(ws, DATA_START_ROW, row)
            ws.cell(row=row, column=1, value=voucher_date_str)
            ws.cell(row=row, column=2, value=entry.voucher_type)
            ws.cell(row=row, column=3, value=voucher_number)
            ws.cell(row=row, column=4, value=line_index)
            ws.cell(row=row, column=5, value=entry.summary if line_index == 1 else "")

            # CRITICAL: 科目编码 must be text format in Excel.
            account_cell = ws.cell(row=row, column=6, value=line.account_code)
            account_cell.number_format = "@"  # Text format

            ws.cell(row=row, column=7, value=line.account_full_name or line.account_name)

            # Debit / Credit amounts
            if line.direction == "debit":
                ws.cell(row=row, column=8, value=float(line.amount))
                ws.cell(row=row, column=9, value="")
            else:
                ws.cell(row=row, column=8, value="")
                ws.cell(row=row, column=9, value=float(line.amount))

            category = _aux_category(line.account_code)
            aux_name = _extract_aux_name(line)
            if not category and entry_counterparty and not line.account_code.startswith(BANK_ACCOUNT_PREFIXES):
                category = "客户" if line.direction == "credit" else "供应商"
                aux_name = entry_counterparty
            if category and aux_name:
                aux_code = _get_aux_code(aux_registry, entry.client_id, category, aux_name)
                registry_changed = True
                if category == "客户":
                    ws.cell(row=row, column=10, value=aux_code)
                    ws.cell(row=row, column=11, value="")
                else:
                    ws.cell(row=row, column=10, value="")
                    ws.cell(row=row, column=11, value=aux_code)
            else:
                ws.cell(row=row, column=10, value="")
                ws.cell(row=row, column=11, value="")

            # Apply basic styles when no template style exists.
            for col in range(1, len(KINGDEE_HEADERS) + 1):
                cell = ws.cell(row=row, column=col)
                if not cell.border or cell.border == Border():
                    cell.border = thin_border
                if col in (8, 9):
                    if not cell.alignment or cell.alignment == Alignment():
                        cell.alignment = amount_alignment
                else:
                    if not cell.alignment or cell.alignment == Alignment():
                        cell.alignment = cell_alignment
                if not cell.font:
                    cell.font = Font(name="微软雅黑", size=9)

            row += 1

    # === Column widths ===
    col_widths = {
        1: 13,   # 日期
        2: 8,    # 凭证字
        3: 10,   # 凭证号
        4: 10,   # 录入顺序
        5: 36,   # 摘要
        6: 15,   # 科目编码
        7: 36,   # 科目名称
        8: 15,   # 借方金额
        9: 15,   # 贷方金额
        10: 10,  # 客户
        11: 10,  # 供应商
    }
    for col, width in col_widths.items():
        letter = get_column_letter(col)
        if not ws.column_dimensions[letter].width:
            ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"

    if registry_changed:
        _save_aux_registry(aux_registry)

    # === Save ===
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"kingdee_vouchers_{ts}.xlsx"
        output_path = settings.export_path / filename
    else:
        output_path = Path(output_path)

    wb.save(output_path)
    logger.info(f"Exported {len(entries)} entries ({row - 2} lines) to {output_path}")

    return output_path
