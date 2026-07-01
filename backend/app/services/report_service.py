"""
Financial report generation from journal entries.

Generates:
- 科目余额表 (Trial Balance): account-level debit/credit totals
- 利润表 (Income Statement): revenue - expenses = profit
"""

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.models.journal_entry import JournalEntry, JournalEntryLine
from app.models.account_subject import AccountSubject


@dataclass
class AccountBalance:
    code: str
    name: str
    category: str
    direction: str
    opening_debit: float = 0.0
    opening_credit: float = 0.0
    current_debit: float = 0.0
    current_credit: float = 0.0
    ending_debit: float = 0.0
    ending_credit: float = 0.0

    @property
    def balance(self) -> float:
        """Net balance: positive=debit, negative=credit."""
        total_debit = self.opening_debit + self.current_debit
        total_credit = self.opening_credit + self.current_credit
        return round(total_debit - total_credit, 2)


def _find_subject(subjects: list[AccountSubject], code: str) -> AccountSubject | None:
    """Find a subject by code, trying exact match then parent match."""
    for s in subjects:
        if s.code == code:
            return s
    # Try parent code (e.g., 5602.01 → 5602)
    parent = ".".join(code.split(".")[:-1]) if "." in code else None
    if parent:
        for s in subjects:
            if s.code == parent:
                return s
    return None


async def get_trial_balance(
    db: AsyncSession,
    client_id: str,
    date_from: date | None = None,
    date_to: date | None = None,
) -> list[dict]:
    """Generate 科目余额表 (Trial Balance) for a client."""
    # Fetch all confirmed entries within date range
    stmt = (
        select(JournalEntry)
        .options(selectinload(JournalEntry.lines))
        .where(JournalEntry.client_id == client_id)
        .where(JournalEntry.status.in_(["confirmed", "exported"]))
    )
    if date_from:
        stmt = stmt.where(JournalEntry.voucher_date >= date_from)
    if date_to:
        stmt = stmt.where(JournalEntry.voucher_date <= date_to)

    stmt = stmt.order_by(JournalEntry.voucher_date)
    result = await db.execute(stmt)
    entries = result.unique().scalars().all()

    # Fetch all subjects for reference
    subj_stmt = select(AccountSubject).where(AccountSubject.is_active == True)
    subj_result = await db.execute(subj_stmt)
    subjects = subj_result.scalars().all()

    # Aggregate by account code
    balances: dict[str, AccountBalance] = {}

    for entry in entries:
        for line in entry.lines:
            code = line.account_code
            if code not in balances:
                subj = _find_subject(subjects, code)
                balances[code] = AccountBalance(
                    code=code,
                    name=line.account_name or code,
                    category=subj.category if subj else "损益",
                    direction=subj.direction if subj else "debit",
                )
            bal = balances[code]
            if line.direction == "debit":
                bal.current_debit += line.amount
            else:
                bal.current_credit += line.amount

    # Calculate ending balances
    for bal in balances.values():
        total_d = round(bal.opening_debit + bal.current_debit, 2)
        total_c = round(bal.opening_credit + bal.current_credit, 2)
        # Determine ending debit/credit based on account direction
        if bal.direction == "debit":
            bal.ending_debit = round(max(total_d - total_c, 0), 2)
            bal.ending_credit = round(max(total_c - total_d, 0), 2)
        else:
            bal.ending_credit = round(max(total_c - total_d, 0), 2)
            bal.ending_debit = round(max(total_d - total_c, 0), 2)

    # Sort by code
    sorted_balances = sorted(balances.values(), key=lambda b: b.code)

    # Calculate totals
    totals = {
        "current_debit": round(sum(b.current_debit for b in sorted_balances), 2),
        "current_credit": round(sum(b.current_credit for b in sorted_balances), 2),
        "ending_debit": round(sum(b.ending_debit for b in sorted_balances), 2),
        "ending_credit": round(sum(b.ending_credit for b in sorted_balances), 2),
    }

    return [
        {
            "code": b.code,
            "name": b.name,
            "category": b.category,
            "direction": b.direction,
            "current_debit": b.current_debit,
            "current_credit": b.current_credit,
            "ending_debit": b.ending_debit,
            "ending_credit": b.ending_credit,
            "balance": b.balance,
        }
        for b in sorted_balances
    ], totals


async def get_income_statement(
    db: AsyncSession,
    client_id: str,
    date_from: date | None = None,
    date_to: date | None = None,
) -> dict:
    """Generate 利润表 (Income Statement) for a client."""
    trial_balance, _ = await get_trial_balance(db, client_id, date_from, date_to)

    # Classify accounts
    revenue_accounts = []       # 收入类 (5001, 5051, 5111, 5301)
    cost_accounts = []           # 成本类 (5401, 5402)
    expense_accounts = []        # 费用类 (5601, 5602, 5603, 5403, 5711, 5801)

    for item in trial_balance:
        code = item["code"]
        if code.startswith("50") or code.startswith("51") or code.startswith("53"):
            # Revenue: credit balance = income
            revenue = item["current_credit"] - item["current_debit"]
            if abs(revenue) > 0.01:
                revenue_accounts.append({**item, "amount": round(revenue, 2)})
        elif code.startswith("54") and code != "5403":
            cost_accounts.append({**item, "amount": round(item["current_debit"] - item["current_credit"], 2)})
        elif code.startswith("5403") or code.startswith("56") or code.startswith("57") or code.startswith("58"):
            expense_accounts.append({**item, "amount": round(item["current_debit"] - item["current_credit"], 2)})

    # Calculate totals
    total_revenue = round(sum(a["amount"] for a in revenue_accounts), 2)
    total_cost = round(sum(a["amount"] for a in cost_accounts), 2)
    gross_profit = round(total_revenue - total_cost, 2)
    total_expense = round(sum(a["amount"] for a in expense_accounts), 2)
    operating_profit = round(gross_profit - total_expense, 2)
    net_profit = operating_profit  # Simplified: no non-operating adjustment here

    return {
        "revenue": revenue_accounts,
        "cost": cost_accounts,
        "expense": expense_accounts,
        "total_revenue": total_revenue,
        "total_cost": total_cost,
        "gross_profit": gross_profit,
        "total_expense": total_expense,
        "operating_profit": operating_profit,
        "net_profit": net_profit,
    }


def export_trial_balance_to_excel(
    trial_balance: list[dict],
    totals: dict,
    output_path: str,
    client_name: str = "",
    date_range: str = "",
):
    """Export 科目余额表 to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "科目余额表"

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value=f"科目余额表 — {client_name}")
    title_cell.font = Font(name="微软雅黑", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    if date_range:
        ws.merge_cells("A2:H2")
        ws.cell(row=2, column=1, value=f"期间: {date_range}").font = Font(name="微软雅黑", size=10)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
        header_row = 3
    else:
        header_row = 2

    # Headers
    headers = ["科目代码", "科目名称", "类别", "方向", "本期借方", "本期贷方", "期末借方", "期末贷方"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for i, item in enumerate(trial_balance):
        row = header_row + 1 + i
        vals = [item["code"], item["name"], item["category"], item["direction"],
                item["current_debit"], item["current_credit"],
                item["ending_debit"], item["ending_credit"]]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = Font(name="微软雅黑", size=9)
            cell.alignment = Alignment(horizontal="right" if col >= 5 else "left")
            cell.border = thin_border
            if col >= 5 and isinstance(v, (int, float)):
                cell.number_format = '#,##0.00'

    # Totals row
    total_row = header_row + 1 + len(trial_balance)
    ws.cell(row=total_row, column=1, value="合计").font = Font(name="微软雅黑", bold=True, size=10)
    for col, key in enumerate(["current_debit", "current_credit", "ending_debit", "ending_credit"], 5):
        cell = ws.cell(row=total_row, column=col, value=totals.get(key, 0))
        cell.font = Font(name="微软雅黑", bold=True, size=10)
        cell.number_format = '#,##0.00'
        cell.border = thin_border

    # Column widths
    widths = [14, 30, 8, 6, 14, 14, 14, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    wb.save(output_path)
    return output_path


def export_income_statement_to_excel(
    data: dict,
    output_path: str,
    client_name: str = "",
    date_range: str = "",
):
    """Export 利润表 to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "利润表"

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:C1")
    ws.cell(row=1, column=1, value=f"利润表 — {client_name}").font = Font(name="微软雅黑", size=14, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    if date_range:
        ws.merge_cells("A2:C2")
        ws.cell(row=2, column=1, value=f"期间: {date_range}").font = Font(name="微软雅黑", size=10)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["项目", "行次", "金额"]
    row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(name="微软雅黑", bold=True, size=10)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center")

    def write_line(r, label, amount, bold=False, indent=0):
        prefix = "  " * indent
        c1 = ws.cell(row=r, column=1, value=prefix + label)
        c1.font = Font(name="微软雅黑", bold=bold, size=10)
        c1.border = thin_border
        c2 = ws.cell(row=r, column=2, value=r - 4)
        c2.border = thin_border
        c3 = ws.cell(row=r, column=3, value=amount if amount else 0)
        c3.font = Font(name="微软雅黑", bold=bold, size=10)
        c3.number_format = '#,##0.00'
        c3.border = thin_border
        c3.alignment = Alignment(horizontal="right")

    r = 5
    # Revenue
    write_line(r, "一、营业收入", data["total_revenue"], bold=True)
    r += 1
    for item in data["revenue"]:
        write_line(r, item["name"], item["amount"], indent=1)
        r += 1

    # Cost
    write_line(r, "二、营业成本", data["total_cost"], bold=True)
    r += 1
    for item in data["cost"]:
        write_line(r, item["name"], item["amount"], indent=1)
        r += 1

    write_line(r, "三、毛利", data["gross_profit"], bold=True)
    r += 1

    # Expenses
    write_line(r, "四、期间费用", data["total_expense"], bold=True)
    r += 1
    for item in data["expense"]:
        write_line(r, item["name"], item["amount"], indent=1)
        r += 1

    write_line(r, "五、营业利润", data["operating_profit"], bold=True)
    r += 1
    write_line(r, "六、净利润", data["net_profit"], bold=True)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18

    wb.save(output_path)
    return output_path
